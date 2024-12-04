#!/usr/bin/python3
##############################################
#
# Name: U7_5_Firewall.py
#
# Author: Peter Christen
#
# Version: 1.3
#
# Date: 20.11.2015
#       22.01.2017 V1.1 Argparse ergänzt
#       27.09.2021 V1.2 workbook-Aufruf angepasst
#       27.09.2024 V1.3 Vollständige Überarbeitung
#
# Purpose: Liest xlsx-Dateien mit Firewall Logs ein
#
##############################################

import sqlite3
import openpyxl
import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors

# Argparse Eingabe prüfen
parser = argparse.ArgumentParser(
    description='Excel Firewall-Log-Auszüge einlesen und neue Excel Files erstellen')
parser.add_argument(
    '-c',
    action='store_true',
    help="Datenbank initial erstellen")
parser.add_argument('-r', metavar='Excel Datei', nargs=1,
                    help="Excel-File mit Firewall-Logauszug einlesen")
parser.add_argument(
    '-p',
    metavar='Port',
    nargs=1,
    help="Excel-File mit Verbindungen für einen spezifischen Port erstellen")
parser.add_argument(
    '-t',
    metavar='IP-Adresse',
    nargs=1,
    help="Excel-File mit Verbindungen für eine spezifische Target-IP erstellen")
parser.add_argument(
    '-s',
    metavar='IP-Adresse',
    nargs=1,
    help="Excel-File mit Verbindungen für eine spezifische Source-IP erstellen")
args = parser.parse_args()

# Grundvariabeln setzen
mydatabase = "U7_5_Firewallog.db"
connection = sqlite3.connect(mydatabase)
cursor = connection.cursor()
exceltargetname = "U7_5_Firewall_Log_Target.xlsx"
excelsourcename = "U7_5_Firewall_Log_Source.xlsx"
excelportname = "U7_5_Firewall_Log_Port.xlsx"

################### Keine Aenderungen mehr nötig ab hier ################

# Functions
def firewall_log_einlesen(sourcefile):
    ''' Excel mit Firewall log einlesen '''
    
    # xlsx-File öffnen
    wb = load_workbook(filename=sourcefile, read_only=True)
    sheet1 = wb.worksheets[0]
    worksheet = wb[sheet1.title]

    # xlsx-File einlesen
    r = 0
    w = []
    for row in worksheet.iter_rows():
        r += 1
        c = 0
        for cell in row:
            c += 1
            w.append(cell.value)

        po = str(w[3]).split(".")
        cursor.execute("replace into firelog values(?,?,?,?,?)",(w[0], w[1], w[2], po[0], w[4]))
        # print w[0],w[1],w[2],po[0],w[4],w[5]
        del w[:]

    cursor.execute("delete from firelog where sourceip like 'Source IP'")
    connection.commit()
    print(str(r) + " Zeilen eingelesen")

def excel_erstellen(typ,muster,ExcelName,Titel):
    ''' Erstelle Excel aus Firewall DB '''

    wb = openpyxl.Workbook()
    ws1 = wb.worksheets[0]
    ws1.title = 'Verbindungen'

    # Styles definieren
    # Font
    fontT = Font(bold=True, size=14, color=colors.BLACK)
    fontb = Font(bold=True, color=colors.BLACK)

    # Alignment
    alignC = Alignment(
        horizontal='center',
        vertical='top',
        text_rotation=0,
        shrink_to_fit=False,
        wrap_text=False)
    alignR = Alignment(
        horizontal='right',
        vertical='top',
        text_rotation=0,
        shrink_to_fit=False,
        wrap_text=False)
    alignL = Alignment(
        horizontal='left',
        vertical='top',
        text_rotation=0,
        shrink_to_fit=False,
        wrap_text=False)

    # Kollonengroesse definieren
    ws1.column_dimensions["A"].width = 14.0
    ws1.column_dimensions["B"].width = 14.0
    ws1.column_dimensions["C"].width = 8.0
    ws1.column_dimensions["D"].width = 7.0

    # Header Zeile schreiben
    ws1['A1'].font = fontT
    ws1['A1'].value = Titel
    ws1['A3'].font = fontb
    ws1['A3'].value = "SourceIP"
    ws1['B3'].font = fontb
    ws1['B3'].value = "TargetIP"
    ws1['C3'].font = fontb
    ws1['C3'].value = "Port"
    ws1['D3'].font = fontb
    ws1['D3'].value = "Protokoll"

    # Daten aus der Datenbanken einfuegen
    z = 4

    if typ == 'target':
        cursor.execute("select * from firelog where targetip=?", (muster,))
    elif typ == 'source':
        cursor.execute("select * from firelog where sourceip=?", (muster,))
    elif typ == 'port':
        cursor.execute("select * from firelog where port=?", (muster,))

    for row in cursor:
        cell=ws1.cell(row=z, column=1,value=row[0])
        cell=ws1.cell(row=z, column=2,value=row[1])
        cell=ws1.cell(row=z, column=3,value=row[2])
        cell=ws1.cell(row=z, column=4,value=row[3])

        z += 1

    wb.save(filename=ExcelName)

# Eingaben auswerten
if args.r:
    firewall_log_einlesen(args.r[0])
elif args.t:
    Titel = "Firewall Verbindungen auf Target IP " + args.t[0]
    excel_erstellen('target',args.t[0],exceltargetname,Titel)
elif args.s:
    Titel = "Firewall Verbindungen von Source IP " + args.s[0]
    excel_erstellen('source',args.s[0],excelsourcename,Titel)
elif args.p:
    Titel = "Firewall Verbindungen auf Port " + args.p[0]
    excel_erstellen('port',args.p[0],excelportname,Titel)
elif args.c:
    cursor.execute(
        'create table if not exists firelog ( sourceip varchar(20), targetip varchar(20), port varchar(20), protokoll varchar(10), count int)')
    cursor.execute(
        'CREATE INDEX if not exists firelog_ind on firelog (sourceip, targetip, port)')
    print("Database created")

connection.commit()
cursor.close()
