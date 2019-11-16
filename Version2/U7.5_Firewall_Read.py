#!/usr/bin/python
# -*- coding: utf-8 -*-
from __future__ import print_function
##############################################
#
# Name: U7.5_Firewall_Read.py
#
# Author: Peter Christen
#
# Version: 1.0
#
# Date: 05.11.2015
#
# Purpose: Liest ein xlsx-File ein und schreibt Daten auf Bildschirm
#
##############################################

import string
import sys
import os
from openpyxl import load_workbook

# Auf Fileangaben prüfen
l = len(sys.argv)
if l == 1:
    sourcefile = 'U7.5_Firewall_Log_Auszug.xlsx'
else:
    sourcefile = sys.argv[1]


# xlsx-File öffnen
wb = load_workbook(filename=sourcefile, read_only=True)
first_sheet = wb.get_sheet_names()[0]
worksheet = wb.get_sheet_by_name(first_sheet)

# xlsx-File einlesen
r = 0
w = []
for row in worksheet.iter_rows():
    r += 1
    c = 0
    for cell in row:
        c += 1
        w.append(cell.value)

    for i in range(len(w)):
        print(w[i], end=' ')

    del w[:]
    print("")

print("\n", r, "Zeilen eingelesen")
