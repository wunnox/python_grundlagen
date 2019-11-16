#!/usr/bin/python
# -*- coding: utf-8 -*-
##############################################
#
# Name: read_tasks.py
#
# Author: Peter Christen
#
# Version: 1.0
# 
# Date: 05.11.2015
#
# Purpose: Liest die TaskListe aus Filemaker ein
#
##############################################

import string, sys, os, openpyxl, pyodbc, sqlite3, datetime, re
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors

pyodbc.pooling = True
connectionF = pyodbc.connect("DRIVER={/Library/ODBC/FileMaker ODBC.bundle/Contents/MacOS/fmodbc.so};SERVER=localhost;DATABASE=Schulung;UID=python;PWD=python12")
cursorF = connectionF.cursor()

connection = sqlite3.connect(':memory:')
#connection = sqlite3.connect("odbc.db")
cursor=connection.cursor()
now = datetime.datetime.now()
today=now.strftime("%Y.%m.%d")
datum=now.strftime("%d.%m.%Y %H:%M")
kw=now.strftime("%V")
start='2015.10.01'
end='2016.12.31'

####### Start Functions
def colsize(col):
         if col>90:
            col=col-26
            b=chr(col)
            col='A'+b
         else:
            col=str(chr(col))
         return col

def da2us(datch):
         d=datch.split(".")
         datus=d[2]+"."+d[1].zfill(2)+"."+d[0].zfill(2)
         return datus

def da_2us(datch):
         d=str(datch).split(" ")
         datus=d[0].replace("-", ".")
         return datus

####### End Functions

####### Datenbanken erstellen
cursor.execute('''create table odbc (
SID varchar(20),
Status varchar(20),
Termin date,
PME date,
PMA date,
PMG date,
Von date,
Bis date,
Zeit varchar(10),
CR varchar(20),
DBA varchar(10),
Size int,
Sort int)''')

cursor.execute('''create table sort ( Termin date primary key)''')

####### Daten aus Filemaker einlesen
r=0
cursorF.execute("select * from DBM")
for row in cursorF:
       von=bis=''
       try:
          zf=row[8].split(" ")
          if zf[0]=="V:":
             von=da2us(zf[1])
             bis=da2us(zf[3])
       except:
           von=bis=''

       #Datumsformat anpassen
       if row[3]!= None:
          termi=da_2us(row[3])
       else:
          termi=""
       try: 
          pme=da_2us(row[12])
       except:
          pme=''
       try: 
          pma=da_2us(row[13])
       except:
          pma=''
       try: 
          pmg=da_2us(row[14])
       except:
          pmg=''

       if row[6]=='Migriert' or row[6]=='Ausgeschaltet':
            so=0
       elif row[6]=='Bereit fuer Migration':
            so=1
       elif row[6]=='In Arbeit':
            so=2
       elif row[6]=='In Vorbereitung':
            so=3
       elif row[6]=='In Planung':
            so=4
       elif row[6]=='In Abklaerung':
            so=5
       elif row[6]=='Offen':
            so=6
       elif row[6]=='Keine Migration':
            so=7
       else:
            so=8

       cursor.execute("replace into odbc values(?,?,?,?,?,?,?,?,?,?,?,?,?)",(row[7],row[6],termi,pme,pma,pmg,von,bis,row[2],row[5],row[4],row[9],so))

####### Migrationstermine auflisten
migdat=[]
cursor.execute("insert or ignore into sort select Termin from odbc")
cursor.execute("insert or ignore into sort select PME from odbc")
cursor.execute("insert or ignore into sort select PMA from odbc")
cursor.execute("insert or ignore into sort select PMG from odbc")
cursor.execute("insert or ignore into sort select Von from odbc")
cursor.execute("insert or ignore into sort select Bis from odbc")
cursor.execute("insert or ignore into sort values(?)",(today,))
cursor.execute("insert or ignore into sort values(?)",(start,))
cursor.execute("insert or ignore into sort values(?)",(end,))
cursor.execute("select Termin from sort where Termin!='' and Termin!=' ' and Termin!='None' order by Termin")
for row in cursor:
        migdat.append(row[0])

days=len(migdat)

from datetime import date
y=now.strftime("%Y")
m=now.strftime("%m")
d=now.strftime("%d")
now = date(int(y), int(m), int(d))

####### Excel erstellen
ExcelName="/Users/tgdchpe1/scripts/Python/Dokumente/DBM Migrationen Planung KW"+kw+".xlsx"
Titel="Planung Datenbank Migrationen Stand "+datum

wb = openpyxl.Workbook()
ws1 = wb.worksheets[0]
ws1.page_setup.orientation = ws1.ORIENTATION_LANDSCAPE
ws1.title = 'Aktuell'

#Styles definieren
#Font
fontT=Font(bold=True,size=14,color=colors.BLACK)
fontb=Font(bold=True,color=colors.BLACK)

#Alignment
alignC=Alignment(horizontal='center',vertical='top',text_rotation=0,shrink_to_fit=False,wrap_text=False)
alignR=Alignment(horizontal='right',vertical='top',text_rotation=0,shrink_to_fit=False,wrap_text=False)
alignL=Alignment(horizontal='left',vertical='top',text_rotation=0,shrink_to_fit=False,wrap_text=False)
alignLW=Alignment(horizontal='left',vertical='top',text_rotation=0,shrink_to_fit=False,wrap_text=True)
align90=Alignment(horizontal='center',vertical='bottom',text_rotation=90,shrink_to_fit=False,wrap_text=False)

#Pattern
fillT = PatternFill(fill_type='solid', start_color='6EB7FF', end_color='6EB7FF')
fillR = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
fillO = PatternFill(fill_type='solid', start_color='FCD020', end_color='FCD020')
fillG = PatternFill(fill_type='solid', start_color='22B604', end_color='22B604')
fillY = PatternFill(fill_type='solid', start_color='FFF251', end_color='FFF251')

#Kollonengroesse definieren
ws1.column_dimensions["A"].width = 10.0
ws1.column_dimensions["B"].width = 14.0
ws1.column_dimensions["C"].width = 8.0
ws1.column_dimensions["D"].width = 7.0
ws1.column_dimensions["E"].width = 1.0
o=5

#Kolonnengroesse Datumsfelder
for i in range(0,days+1):
   col=65+o+i
   col=colsize(col)
   ws1.column_dimensions[col].width = 3.0
   ce=col+str(3)
   ws1[ce].font=Font(bold=True,color=colors.BLACK)
   #print col

col=65+o+i+1
col=colsize(col)
ce=col+str(2)
legende=col
ws1.column_dimensions[col].width = 28.0
ws1[ce].font=fontb
ws1[ce].value="Legende"
ce=col+str(3)
ws1[ce].alignment=alignLW
ws1[ce].value="PE=Portmatrix Request eingereicht\nPA=Portmatrix aktiviert\nPG=Portmatrix geprueft\nOrange: Migration abgemacht\nGelb: Migration geplant"

ws1['A3'].font=fontb
ws1['A3'].value="SID"
ws1['B3'].font=fontb
ws1['B3'].value="Status"
ws1['C3'].font=fontb
ws1['C3'].value="DBA"
ws1['D3'].font=fontb
ws1['D3'].value="Size GB"

#Datumsfelder einfuellen
cpos={}
coln=65+5
for i in migdat:
    col=coln
    col=colsize(col)
    ce=col+str(3)
    i2=i.split(".")
    i3=i2[2]+"."+i2[1]+"."+i2[0]
    ws1[ce].font=fontb
    if i==today:
       ws1[ce].fill=fillT
       ws1[ce].alignment=align90
       ws1[ce].value=i3
       cev=coln-1
       ceb=coln+1
       cev=colsize(cev)+str(2)
       ceb=colsize(ceb)+str(2)
       cevb=cev+':'+ceb
       ws1.merge_cells(cevb)
       ws1[cev].fill=fillT
       ws1[cev].alignment=alignC
       ws1[cev].font=fontb
       ws1[cev].value="Heute"
    else:
       ws1[ce].alignment=align90
       ws1[ce].value=i3
    cpos[i]=coln
    coln+=1

#Titelzeile einfüllen
col=colsize(coln+2)
post='A1:'+str(col)+'1'
ws1.merge_cells(post)
ws1['A1'].fill=fillT
ws1['A1'].font=fontT
ws1['A1']=Titel

#Daten der Datenbanken einfuegen
coln=65
z=4
heute = datetime.datetime.now()
dba=[]
dbaname="None"
cursor.execute("select * from odbc where Status != 'Keine Migration' order by Sort,SID")
for row in cursor:
    col=colsize(coln)
    ce=col+str(z)
    ws1[ce].value=row[0]
    col=colsize(coln+1)
    ce=col+str(z)
    ws1[ce].value=row[1]
    col=colsize(coln+2)
    ce=col+str(z)
    ws1[ce].value=row[10]
    col=colsize(coln+3)
    ce=col+str(z)
    ws1[ce].value=row[11]

    #Portmatrix Positionen
    if row[3] != "" and row[3] != "None":
       pmepos=row[3]
       col=colsize(cpos[pmepos])
       ce=col+str(z)
       ws1[ce].value="PE"

    if row[4] != "" and row[4] != "None":
       pmapos=row[4]
       col=colsize(cpos[pmapos])
       ce=col+str(z)
       ws1[ce].value="PA"

    if row[5] != "" and row[5] != "None":
       pmgpos=row[5]
       col=colsize(cpos[pmgpos])
       ce=col+str(z)
       ws1[ce].value="PG"

    if row[12]>1:
       if row[6]!='':
           vonpos=row[6]
           colv=colsize(cpos[vonpos])
           ce=colv+str(z)
           ws1[ce].fill=fillT
           ws1[ce].value="|->"
       if row[7]!='':
           bispos=row[7]
           colb=colsize(cpos[bispos])
           ce=colb+str(z)
           ws1[ce].fill=fillT
           ws1[ce].value="<-|"
           if cpos[bispos]-cpos[vonpos]>1:
              for colr in range(cpos[vonpos]+1,cpos[bispos]):
                 col=colsize(colr)
                 ce=col+str(z)
                 ws1[ce].fill=fillT
                 ws1[ce].value=" "

    if row[2] != "":
       migposd=row[2].split(".")
       ja=int(migposd[0])
       mo=int(migposd[1])
       ta=int(migposd[2])
       migpos=row[2]
       col=colsize(cpos[migpos])
       ce=col+str(z)
       if row[1]=="Migriert" or row[1]=="Ausgeschaltet":
          ws1[ce].fill=fillG
          ws1[ce].value=" "
       elif (row[1]=="In Arbeit" or row[1]=="Bereit fuer Migration") and row[2] != end:
          diff=(heute-datetime.datetime(ja,mo,ta,16,0)).total_seconds()
          if diff>0:
             ws1[ce].fill=fillR
             ws1[ce].value=" "
          else:
             ws1[ce].fill=fillO
             ws1[ce].value=" "
       else:
          if row[2] != end:
             ws1[ce].fill=fillY
             ws1[ce].value=" "

       if row[8]!=None:
          col=colsize(cpos[migpos]+1)
          ce=col+str(z)
          ws1[ce].value=row[8]

    ce=legende+str(z)
    ws1[ce].value=row[9]

    z+=1

connection.commit()
wb.save(filename = ExcelName)
cursor.close()
cursorF.close()
quit()

