##############################################
#
# Name: U7_5_Firewall_Read.py
#
# Author: Peter Christen
#
# Version: 1.2
#
# Date: 05.11.2015 V1.0
#       27.09.2021 V1.1 workbook-Aufruf angepasst
#       10.10.2023 V1.2 Script vereinfacht
#
# Purpose: Liest ein xlsx-File ein und schreibt Daten auf Bildschirm
#
##############################################

import sys
from openpyxl import load_workbook

# Auf Fileangaben prüfen
l = len(sys.argv)
if l == 1:
    sourcefile = 'U7_5_Firewall_Log_Auszug.xlsx'
else:
    sourcefile = sys.argv[1]

# xlsx-File öffnen
wb = load_workbook(filename=sourcefile, read_only=True)
worksheet = wb.worksheets[0]

# xlsx-File einlesen
r = 0
for row in worksheet.iter_rows():
    r += 1
    for cell in row:
        print(cell.value, end='\t')
    print()

print(f"\n{r} Zeilen eingelesen")
